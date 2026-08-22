"""Parser for Salt Lake County's 2026 primary summary PDF, which uses a
distinct format:

    Voters Cast: 103,419 of 265,746 (38.92%)
    U.S. HOUSE DISTRICT 1 (DEM) (Vote for 1)
    (DEM)
                        Total
    Times Cast          57,455 / 103,618   55.45%
    Candidate           Total
    BEN MCADAMS       29,737        51.90%
    ...
    Total Votes       57,295

No explicit Over/Under votes; turnout statistics are county-wide totals only.
Salt Lake's precinct-level data is an .xlsx (handled separately).
"""

import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
import p2026_common as C

_PARTY_CODES = {"REP", "DEM", "LIB", "IAP", "CON", "UUP", "IND", "GRN", "NP"}
_VOTERS_CAST_RE = re.compile(r"Voters\s*Cast:\s*([\d,]+)\s+of\s+([\d,]+)\s+\([\d.]+%\)", re.IGNORECASE)
_HEADER_RE = re.compile(r"^(?P<office>.+?)(?:\s*\((?P<party>REP|DEM|LIB|IAP|CON|UUP|IND|GRN|NP)\))?\s*\(Vote\s+for\s+\d+\)\s*$", re.IGNORECASE)
_CAND_RE = re.compile(r"^(?P<name>.+?)\s+(?P<votes>[\d,]+)\s+[\d.]+%\s*$")
_NUM_ONLY_RE = re.compile(r"^(?P<name>.+?)\s+(?P<votes>[\d,]+)\s*$")

_SKIP_PREFIXES = ("TIMES CAST", "TOTAL VOTES", "CANDIDATE", "PAGE", "FINAL",
                   "2026 PRIMARY", "TUESDAY", "SALT LAKE", "VOTE BY MAIL")
_SKIP_EXACT = {"TOTAL", "CANDIDATE", "TOTAL VOTES"}


def _is_skip(text):
    up = text.strip().upper()
    if up in _SKIP_EXACT:
        return True
    return up.startswith(_SKIP_PREFIXES)


def parse_summary(county, pdf_path):
    pages = C.extract_rows(pdf_path, county)
    display = C.COUNTY_DISPLAY.get(county, county)
    rows = []
    office, district, party = "", "", ""
    for page in pages:
        for _, _, text in page:
            t = text.strip()
            if not t:
                continue
            # County-wide turnout line -> Registered Voters / Ballots Cast
            vm = _VOTERS_CAST_RE.search(t)
            if vm:
                rows.append(C.meta_row(display, C.BALLOTS_CAST, "", vm.group(1)))
                rows.append(C.meta_row(display, C.REGISTERED_VOTERS, "", vm.group(2)))
                continue
            if _is_skip(t):
                continue
            # "(DEM)" standalone line -> skip
            if t.strip("()").upper() in _PARTY_CODES:
                continue
            hdr = _HEADER_RE.match(t)
            if hdr:
                office_text = hdr.group("office").strip()
                office = C.normalize_office(office_text)
                district = C.parse_district(office_text)
                party = (hdr.group("party") or "").upper()
                continue
            # Candidate line with percentage
            cm = _CAND_RE.match(t)
            if not cm:
                cm = _NUM_ONLY_RE.match(t)
            if cm:
                name = cm.group("name").strip()
                if _is_skip(name):
                    continue
                rows.append(C.candidate_row(display, office, district, party,
                                             C.title_case_name(name), cm.group("votes")))
                continue
    return rows


# --------------------------------------------------------------------------- #
# Precinct parser (Statement of Votes Cast .xlsx)
# --------------------------------------------------------------------------- #
#
# Layout: one workbook per county, 17 sheets.
#   Sheet1  — turnout: Precinct | Registered Voters | Voters Cast | % Turnout
#   Sheet2+ — one contest per sheet.  Row 1 = contest title
#             ("U.S. HOUSE DISTRICT 1 (DEM) (Vote for 1)"); row 3 = column
#             header.  Columns: Precinct | Times Cast | Registered Voters |
#             [blank] | Precinct | <candidate columns ...> | Total Votes |
#             Unresolved Write-In.  Candidate header cells carry the name plus
#             a party marker ("BEN MCADAMS\n((DEM)) ").  Suppressed cells
#             (voter-privacy) show "****" and are skipped — their votes are
#             counted in the summary totals but hidden per-precinct, so precinct
#             sums run slightly below summary totals (same suppression pattern
#             as the scanned counties).

import openpyxl

_XLSX_TITLE_RE = re.compile(
    r"^(?P<office>.+?)(?:\s*\((?P<party>REP|DEM|LIB|IAP|CON|UUP|IND|GRN|NP)\))?"
    r"\s*\(Vote\s+for\s+\d+\)", re.IGNORECASE)

# header labels that are NOT candidate columns
_NON_CAND_LABELS = {
    "total votes", "unresolved\nwrite-in", "unresolved write-in",
    "unresolved\nwrite-in ", "total", "write-in", "times cast",
    "registered\nvoters", "registered voters", "% turnout", "precinct",
}

# Salt Lake precinct codes: 2+ uppercase letters then a digit (ALT001, SLC050,
# WVC903, ...).  Matches real precinct codes; rejects 'County - Total',
# 'Salt Lake County - Total', 'Cumulative', etc.
_PRECINCT_CODE_RE = re.compile(r"^[A-Z]{2,}\d")


def _int_cell(v):
    """Parse an integer from a cell; None/blank/'****' -> None."""
    if v is None:
        return None
    s = str(v).strip().replace(",", "")
    if s in ("", "****", "None", "%"):
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def _title_parts(title):
    """office, district, party from a contest sheet title."""
    first = title.splitlines()[0].strip()
    m = _XLSX_TITLE_RE.match(first)
    if not m:
        return None
    office_text = m.group("office").strip()
    return (C.normalize_office(office_text),
            C.parse_district(office_text),
            (m.group("party") or "").upper())


def _is_precinct_code(v):
    """True only for a real precinct code (e.g. 'ALT001').  A positive regex
    keeps out the county-total / cumulative rows ('County - Total',
    'Salt Lake County - Total', 'Cumulative', ...) whose votes would otherwise
    be double-counted against the summary."""
    if v is None:
        return False
    return bool(_PRECINCT_CODE_RE.match(str(v).strip()))


def parse_precinct(county, xlsx_path):
    """Parse Salt Lake's Statement of Votes Cast .xlsx into precinct rows."""
    display = C.COUNTY_DISPLAY.get(county, county)
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheets = wb.sheetnames
    rows = []

    # --- Sheet1: per-precinct Registered Voters / Ballots Cast --- #
    # Columns are header-driven (the sheet has blank spacer columns, so the
    # "Voters Cast" column is not always at a fixed index).
    if sheets:
        data = list(wb[sheets[0]].iter_rows(values_only=True))
        hdr_idx = None
        for ri, r in enumerate(data):
            if r and str(r[0]).strip().lower() == "precinct":
                hdr_idx = ri
                break
        if hdr_idx is not None:
            reg_i = vc_i = None
            for i, c in enumerate(data[hdr_idx]):
                if c is None:
                    continue
                label = str(c).splitlines()[0].strip().lower()
                if reg_i is None and label.startswith("registered"):
                    reg_i = i
                elif vc_i is None and ("voters cast" in label or "ballots cast" in label):
                    vc_i = i
            for r in data[hdr_idx + 1:]:
                if not r or not _is_precinct_code(r[0]):
                    continue
                code = str(r[0]).strip()
                reg = _int_cell(r[reg_i]) if reg_i is not None and reg_i < len(r) else None
                vc = _int_cell(r[vc_i]) if vc_i is not None and vc_i < len(r) else None
                if reg is not None:
                    rows.append(C.meta_row(display, C.REGISTERED_VOTERS,
                                            "", reg, precinct=code))
                if vc is not None:
                    rows.append(C.meta_row(display, C.BALLOTS_CAST,
                                            "", vc, precinct=code))

    # --- Contest sheets (Sheet2+): candidate votes per precinct --- #
    for sn in sheets[1:]:
        data = list(wb[sn].iter_rows(values_only=True))
        if len(data) < 2 or data[1][0] is None:
            continue
        tp = _title_parts(str(data[1][0]))
        if tp is None:
            continue
        office, district, party = tp

        # locate the column-header row (first row whose col0 == 'Precinct')
        hdr_idx = None
        for ri, r in enumerate(data):
            if r and str(r[0]).strip().lower() == "precinct":
                hdr_idx = ri
                break
        if hdr_idx is None:
            continue
        hdr = data[hdr_idx]

        # candidate columns start after the second 'Precinct' column
        pcols = [i for i, c in enumerate(hdr)
                 if c is not None and str(c).strip().lower() == "precinct"]
        start = pcols[1] + 1 if len(pcols) >= 2 else pcols[0] + 1
        cand_cols = []  # (col_index, candidate_name)
        for i in range(start, len(hdr)):
            c = hdr[i]
            if c is None:
                continue
            name = str(c).splitlines()[0].strip()
            if not name or name.lower() in _NON_CAND_LABELS:
                continue
            cand_cols.append((i, name))

        for r in data[hdr_idx + 1:]:
            if not r or not _is_precinct_code(r[0]):
                continue
            code = str(r[0]).strip()
            for ci, name in cand_cols:
                v = _int_cell(r[ci]) if ci < len(r) else None
                if v is None:
                    continue  # suppressed (****) or blank -> omit
                rows.append(C.candidate_row(display, office, district, party,
                                            C.title_case_name(name), v,
                                            precinct=code))
    return rows


if __name__ == "__main__":
    src = C.find_source("salt_lake", "summary")
    rows = parse_summary("salt_lake", src)
    ccsv, _ = C.county_paths("salt_lake")
    C.write_csv(rows, ccsv, precinct=False)
    issues = C.sanity_check(ccsv, precinct=False)
    print(f"{ccsv.name}: {len(rows)} rows, issues={issues}")
    for r in rows[:12]:
        print(" ", r)